#! /usr/bin/env python
__author__ = "Willy Fernandez"
__author_email__ = "wilferna@cisco.com"
__copyright__ = "Copyright (c) 2022 Cisco Systems, Inc."
__license__ = "MIT"
"""
Script to generate a summary report base on data source generate by Matrix
"""

import openpyxl

def buildHeaders(sh3):
    sh3['D3'].value = "April's PCRF Monthly KPI Report Sundance Network"
    sh3['B4'].value = "Site"
    sh3['C4'].value = "KPI"
    sh3['D4'].value = "VM"
    sh3['E4'].value = "Index"
    sh3['F4'].value = "Critical"
    sh3['G4'].value = "Threshold"
    sh3['H4'].value = "Min"
    sh3['I4'].value = "Max"
    sh3['J4'].value = "Avg"
    return sh3

book = openpyxl.load_workbook('ATT_PCRF_SD.xlsx') #Workbook object
sh1 = book['Summary2']
rows = sh1.rows #list of Cell elements

sh3 = buildHeaders(book.create_sheet(0))
sh3.title="SummaryFinal"
count = 7
for cell in rows:
    if cell[2].value =="memory" and cell[9].value > 80:
        vm = cell[1].value
        percentage = cell[9].value
        sh3.cell(row=count, column=2, value=vm)
        sh3.cell(row=count, column=3, value=percentage)
        count = count + 1
book.save('FinalReport.xlsx')
print('Report generated')
