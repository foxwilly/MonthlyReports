#! /usr/bin/env python
__author__ = "Willy Fernandez"
__author_email__ = "wilferna@cisco.com"
__copyright__ = "Copyright (c) 2022 Cisco Systems, Inc."
__license__ = "MIT"
"""
Script to generate a summary report base on data source generate by Matrix
"""

import openpyxl
from openpyxl.styles import Font, Border, Side, NamedStyle, PatternFill
from datetime import date
import re

''' to include titles in the sheet '''
def buildHeaders(sh3,rownumb, named_style, rack):
    sh3.cell(row=rownumb-1, column=2, value="April's PCRF Monthly KPI Report Sundance  "+rack).font = Font(name="Arial", size=14, bold=True)
    sh3.cell(row=rownumb, column=2, value="Site").style = named_style
    sh3.cell(row=rownumb, column=3, value="KPI").style = named_style
    sh3.cell(row=rownumb, column=4, value="VM").style = named_style
    sh3.cell(row=rownumb, column=5, value="Index").style = named_style
    sh3.cell(row=rownumb, column=6, value="Critical").style = named_style
    sh3.cell(row=rownumb, column=7, value="Threshold").style = named_style
    sh3.cell(row=rownumb, column=8, value="Min").style = named_style
    sh3.cell(row=rownumb, column=9, value="Max").style = named_style
    sh3.cell(row=rownumb, column=10, value="Avg").style = named_style
    return sh3


''' to include yellow line to separate the kpis '''
def printYellowLine(sh3, count):
    sh3.cell(row=count, column=2).fill=PatternFill("solid",fgColor="00FFFF00")
    sh3.cell(row=count, column=3).fill=PatternFill("solid",fgColor="00FFFF00")
    sh3.cell(row=count, column=4).fill=PatternFill("solid",fgColor="00FFFF00")
    sh3.cell(row=count, column=5).fill=PatternFill("solid",fgColor="00FFFF00")
    sh3.cell(row=count, column=6).fill=PatternFill("solid",fgColor="00FFFF00")
    sh3.cell(row=count, column=7).fill=PatternFill("solid",fgColor="00FFFF00")
    sh3.cell(row=count, column=8).fill=PatternFill("solid",fgColor="00FFFF00")
    sh3.cell(row=count, column=9).fill=PatternFill("solid",fgColor="00FFFF00")
    sh3.cell(row=count, column=10).fill = PatternFill("solid", fgColor="00FFFF00")
    return sh3

def printExecutiveSum(sh3, count, result):
    sh3.cell(row=count, column=40, value=result['site']).fill = PatternFill("solid", fgColor="00FFFF00")
    sh3.cell(row=count, column=41, value=result['kpi']).fill = PatternFill("solid", fgColor="00FFFF00")
    sh3.cell(row=count, column=42, value=result['lb']).fill=PatternFill("solid",fgColor="003FCF40")
    sh3.cell(row=count, column=43, value=result['qns']).fill=PatternFill("solid",fgColor="003FCF40")
    sh3.cell(row=count, column=44, value=result['cc']).fill=PatternFill("solid",fgColor="003FCF40")
    sh3.cell(row=count, column=45, value=result['udc']).fill=PatternFill("solid",fgColor="003FCF40")
    sh3.cell(row=count, column=46, value=result['lwr']).fill=PatternFill("solid",fgColor="003FCF40")
    sh3.cell(row=count, column=47, value=result['gxCCRI']).fill=PatternFill("solid",fgColor="003FCF40")
    sh3.cell(row=count, column=48, value=result['gxCCRU']).fill=PatternFill("solid",fgColor="003FCF40")
    sh3.cell(row=count, column=49, value=result['gxCCRT']).fill=PatternFill("solid",fgColor="003FCF40")
    sh3.cell(row=count, column=50, value=result['gx']).fill=PatternFill("solid",fgColor="003FCF40")
    sh3.cell(row=count, column=51, value=result['sy']).fill=PatternFill("solid",fgColor="003FCF40")
    sh3.cell(row=count, column=52, value=result['udcs']).fill=PatternFill("solid",fgColor="003FCF40")
    sh3.cell(row=count, column=53, value=result['unknown']).fill=PatternFill("solid",fgColor="003FCF40")
    return sh3


''' 
to read the offending data from excel file
 @mthreshold  threshold for memory kpi
 @cputhreshold  threshold for the cpi kpi
 '''
def loadData(mthreshold, cputhreshold):
    book = openpyxl.load_workbook('ATT-ATT_Monthly_Worksheet.xlsx') #Workbook object
    sh1 = book['Summary']
    rows = sh1.rows #list of Cell elements
    data = dict()
    sum_row = [] #list with data to print in summary sheet
    for cell in rows:
        if cell[2].value == "memory" and cell[9].value > mthreshold:  # for CPS Memory Usage
            site = (cell[0].value[0:7])
            vm = cell[1].value
            kpi = cell[3].value
            index = cell[5].value
            min = cell[9].value
            max = cell[10].value
            avg = cell[11].value
            # populate
            data = {'site': site, 'vm': vm, 'kpi': kpi, 'index': index, 'threshold': mthreshold, 'val': [min, max, avg]}
            sum_row.append(data)
        else:
            if cell[2].value == "cpu" and cell[10].value > cputhreshold:  # for CPU Usage - Per VM

                site = (cell[0].value[0:7])
                vm = cell[1].value
                kpi = cell[3].value
                index = cell[5].value
                min = cell[9].value
                max = cell[10].value
                avg = cell[11].value
                data = {'site': site,  'vm': vm, 'kpi': kpi, 'index': index, 'threshold': cputhreshold, 'val': [min, max, avg]}
                sum_row.append(data)
            else:
                if cell[2].value == "node.counters" and re.match('Gx_CCR-[I|U|T]_[5|3]',
                                                                 cell[5].value):  # for DIAMETER ERRORS
                    site = (cell[0].value[0:7])
                    vm = cell[1].value
                    kpi = cell[3].value
                    index = cell[5].value
                    max = cell[10].value
                    data = {'site': site, 'vm': vm, 'kpi': kpi, 'index': index, 'threshold': '', 'val': [0, max, 0]}
                    sum_row.append(data)
                else:
                    if cell[3].value == "CPS Session Summary Index" and re.search('cc01v', cell[1].value ): # for SESSIONS, on from cc01, cc02 has same info
                        site = (cell[0].value[0:7])
                        vm = cell[1].value
                        kpi = cell[3].value
                        index = cell[5].value
                        min = cell[9].value
                        max = cell[10].value
                        avg = cell[11].value
                        # to populate the dict
                        data = {'site': site, 'vm': vm, 'kpi': kpi, 'index': index, 'threshold': '', 'val': [min, max, avg]}
                        sum_row.append(data)
    newsh = book.create_sheet(0) # to create the new sheet
    newsh.title = "Monthly Report"
    font = Font(bold=True, size=11)
    thick = Side(style="thick")
    border = Border(left=thick, right=thick, top=thick, bottom=thick)
    named_style = NamedStyle(name="highlightx", font=font, border=border)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    named_style2 = NamedStyle(name="highlight3", border=border)
    sum_row = sorted(sum_row, key=lambda x: x['kpi'])
    shes = book['Executive Summary'] # Excecutive Summary
    nrorow = createSummarySheet(newsh, sum_row , 3, "akrnoh1", named_style, named_style2)
    print(createExecutiveSum(shes, sum_row, 8, "akrnoh1", "CPS Memory Usage"))
    print(createExecutiveSum(shes, sum_row, 9, "akrnoh1", "CPU Usage - Per VM"))
    print(createExecutiveSum(shes, sum_row, 10, "akrnoh1", "CPS Session Summary Index"))
    nrorow = createSummarySheet(newsh, sum_row, nrorow + 2, "allntx1", named_style, named_style2)
    print(createExecutiveSum(shes, sum_row, 31, "allntx1", "CPS Memory Usage"))
    print(createExecutiveSum(shes, sum_row, 32, "allntx1", "CPU Usage - Per VM"))
    print(createExecutiveSum(shes, sum_row, 33, "allntx1", "CPS Session Summary Index"))
    nrorow = createSummarySheet(newsh, sum_row, nrorow + 2, "alpsga1", named_style, named_style2)
    print(createExecutiveSum(shes, sum_row, 54, "alpsga1", "CPS Memory Usage"))
    print(createExecutiveSum(shes, sum_row, 55, "alpsga1", "CPU Usage - Per VM"))
    print(createExecutiveSum(shes, sum_row, 56, "alpsga1", "CPS Session Summary Index"))
    nrorow = createSummarySheet(newsh, sum_row, nrorow + 2, "artnva1", named_style, named_style2)
    print(createExecutiveSum(shes, sum_row, 77, "artnva1", "CPS Memory Usage"))
    print(createExecutiveSum(shes, sum_row, 78, "artnva1", "CPU Usage - Per VM"))
    print(createExecutiveSum(shes, sum_row, 79, "artnva1", "CPS Session Summary Index"))
    nrorow = createSummarySheet(newsh, sum_row, nrorow + 2, "bothwa1", named_style, named_style2)
    print(createExecutiveSum(shes, sum_row, 100, "bothwa1", "CPS Memory Usage"))
    print(createExecutiveSum(shes, sum_row, 101, "bothwa1", "CPU Usage - Per VM"))
    print(createExecutiveSum(shes, sum_row, 102, "bothwa1", "CPS Session Summary Index"))
    nrorow = createSummarySheet(newsh, sum_row, nrorow + 2, "chcgil1", named_style, named_style2)
    print(createExecutiveSum(shes, sum_row, 123, "chcgil1", "CPS Memory Usage"))
    print(createExecutiveSum(shes, sum_row, 124, "chcgil1", "CPU Usage - Per VM"))
    print(createExecutiveSum(shes, sum_row, 125, "chcgil1", "CPS Session Summary Index"))
    nrorow = createSummarySheet(newsh, sum_row, nrorow + 2, "cncrca1", named_style, named_style2)
    print(createExecutiveSum(shes, sum_row, 146, "cncrca1", "CPS Memory Usage"))
    print(createExecutiveSum(shes, sum_row, 147, "cncrca1", "CPU Usage - Per VM"))
    print(createExecutiveSum(shes, sum_row, 148, "cncrca1", "CPS Session Summary Index"))
    nrorow = createSummarySheet(newsh, sum_row, nrorow + 2, "gsvlfl1", named_style, named_style2)
    print(createExecutiveSum(shes, sum_row, 169, "gsvlfl1", "CPS Memory Usage"))
    print(createExecutiveSum(shes, sum_row, 170, "gsvlfl1", "CPU Usage - Per VM"))
    print(createExecutiveSum(shes, sum_row, 171, "gsvlfl1", "CPS Session Summary Index"))
    nrorow = createSummarySheet(newsh, sum_row, nrorow + 2, "hstntx1", named_style, named_style2)
    print(createExecutiveSum(shes, sum_row, 192, "hstntx1", "CPS Memory Usage"))
    print(createExecutiveSum(shes, sum_row, 193, "hstntx1", "CPU Usage - Per VM"))
    print(createExecutiveSum(shes, sum_row, 194, "hstntx1", "CPS Session Summary Index"))
    nrorow = createSummarySheet(newsh, sum_row, nrorow + 2, "nycmny1", named_style, named_style2)
    print(createExecutiveSum(shes, sum_row, 215, "nycmny1", "CPS Memory Usage"))
    print(createExecutiveSum(shes, sum_row, 216, "nycmny1", "CPU Usage - Per VM"))
    print(createExecutiveSum(shes, sum_row, 217, "nycmny1", "CPS Session Summary Index"))
    nrorow = createSummarySheet(newsh, sum_row, nrorow + 2, "vnnyca1", named_style, named_style2)
    print(createExecutiveSum(shes, sum_row, 238, "vnnyca1", "CPS Memory Usage"))
    print(createExecutiveSum(shes, sum_row, 239, "vnnyca1", "CPU Usage - Per VM"))
    print(createExecutiveSum(shes, sum_row, 240, "vnnyca1", "CPS Session Summary Index"))
    book.save('ATT_Monthly_Report.xlsx')


''' create a sheet and fill the data base on the list sum_row
@book  excel files where the new sheet is
@sum_row  list with the populated data to print in the summary sheet
'''
def createSummarySheet(newsh, sum_row, nrorow, rack, named_style, named_style2):
    data1 = {'site': '', 'kpi': '', 'vm': '', 'threshold': '', 'val': ''}
    data2 = {'site': '', 'kpi': '', 'vm': '', 'threshold': '', 'val': ''}
    sh3 = buildHeaders(newsh, nrorow, named_style, rack)
    #sh3.title = "SummaryFinal"
    rowCount = nrorow+2
    acum_session = 0
    acum_counter = 0
    yellowLine = ""
    for element in sum_row:
        if element['site'] == rack:
            if element['kpi'] == "CPS Memory Usage":
                if yellowLine != "CPS Memory Usage":
                    sh3 = printYellowLine(sh3, rowCount)
                    rowCount = rowCount + 1
                    yellowLine = "CPS Memory Usage"
                sh3 = printRow(sh3, rowCount, element, named_style2, False)
                rowCount = rowCount + 1
            else:
                if element['kpi'] == "CPU Usage - Per VM":
                    if yellowLine != "CPU Usage - Per VM":
                        sh3 = printYellowLine(sh3, rowCount)
                        rowCount = rowCount + 1
                        yellowLine = "CPU Usage - Per VM"
                    sh3 = printRow(sh3, rowCount, element, named_style2, False)
                    rowCount = rowCount + 1
                else:
                    if element['kpi'] == "CPS QNS Counters":
                        acum_counter = element['val'][1] + acum_counter
                        data1 = {'site': element['site'], 'vm': 'CCR I/U/T', 'kpi': element['kpi'], 'index': element['index'], 'threshold': element['threshold'], 'val': ["NA", acum_counter, "NA"]}
                    else:
                        if element['kpi'] == "CPS Session Summary Index":
                            acum_session = element['val'][1] + acum_session
                            data2 = {'site': element['site'], 'vm': element['vm'], 'kpi': element['kpi'], 'index': element['index'], 'threshold': element['threshold'],
                                     'val': ["NA", acum_session, "NA"]}
                            if yellowLine != "CPS Session Summary Index":
                                sh3 = printYellowLine(sh3, rowCount)
                                rowCount = rowCount + 1
                                yellowLine = "CPS Session Summary Index"
                            sh3 = printRow(sh3, rowCount, element, named_style2,True)
                            rowCount = rowCount + 1
    if data1['site'] == rack:
        sh3 = printYellowLine(sh3, rowCount)
        rowCount = rowCount + 1
        sh3 = printRow(sh3, rowCount, data1, named_style2, False)
        rowCount = rowCount + 1

   # if data2['site'] == rack:
   #     sh3 = printYellowLine(sh3, rowCount)
   #     rowCount = rowCount + 1
   #     sh3 = printRow(sh3, rowCount, data2, named_style2)
   #     rowCount = rowCount + 1
    #book.save('FinalReport.xlsx')
    return rowCount

def printRow(sh3, count, dict, named_style2, index):
    sh3.cell(row=count, column=2, value=dict['site']).style = named_style2
    sh3.cell(row=count, column=3, value=dict['kpi']).style = named_style2
    sh3.cell(row=count, column=4, value=dict['vm']).style = named_style2
    if index:
        sh3.cell(row=count, column=5, value=dict['index']).style = named_style2
    else:
        sh3.cell(row=count, column=5).style = named_style2
    sh3.cell(row=count, column=6).style = named_style2
    sh3.cell(row=count, column=7, value=str(dict['threshold'])+'%').style = named_style2
    sh3.cell(row=count, column=8, value=dict['val'][0]).style = named_style2
    sh3.cell(row=count, column=9, value=dict['val'][1]).style = named_style2
    sh3.cell(row=count, column=10, value=dict['val'][2]).style = named_style2
    return sh3

def createExecutiveSum(newsh, sum_row, nrorow, rack, kpi):
    lb = 0
    qns = 0
    cc = 0
    udc = 0 #for udc vm
    lwr = 0
    gxCCRI = 0
    gxCCRU = 0
    gxCCRT = 0
    gx = 0
    sy = 0
    udcs = 0 #for udc sessions
    unknown =0
    result = dict()
    for element in sum_row:
        if element['site'] == rack:
            if element['kpi'] == kpi:
                if re.search('lb0', element['vm']):
                    lb = lb+1
                elif re.search('ps[m01]', element['vm']):
                    qns = qns+1
                elif re.search('cc0', element['vm']):
                    cc = cc+1
                elif re.search('ud0', element['vm']):
                    udc = udc+1
                elif re.search('lwr0', element['vm']):
                    lwr = lwr+1
                else:
                    unknown =unknown+1
                    print(element['vm'])
                if element['index'] == "GX_TGPP":
                    gx = element['val'][2]
                if element['index'] == "SY_V11":
                    sy = element['val'][2]
                if element['index'] == "UDC_FE":
                    udcs = element['val'][2]

    result = {'site': rack, 'kpi': kpi, 'lb': lb, 'qns': qns, 'cc': cc, 'udc': udc, 'lwr': lwr, 'gxCCRI': gxCCRI, 'gxCCRU': gxCCRU, 'gxCCRT': gxCCRT, 'gx': gx, 'sy': sy, 'udcs': udcs, 'unknown': unknown}
    printExecutiveSum(newsh, nrorow, result)
    return result


print(date.today())
loadData(80, 60)
print('Report generated')
